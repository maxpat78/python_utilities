"Esporta in un file json Comune (o Stato estero), sigla automobilistica e codice catastale da un file .cfg delle app SOGEI e dai fogli XSLX dell'ISTAT"
from openpyxl import load_workbook
import json

territori = {} # {CODICE_CAT: (nome, sigla)}

print ("Carico l'elenco di Comuni (attuali e non)...")
with open("ListaComuni.cfg", "r", encoding="utf-8") as f:
    for line in f:
        O = line.split(';')
        if len(O[0]) != 4: continue # la prima riga non ha un cod. catastale
        territori[O[0]] = (O[1].upper(), O[2])

print ('Carico gli Stati esteri...')
wb = load_workbook('Elenco-codici-e-denominazioni-al-31_12_2023.xlsx') # https://www.istat.it/classificazione/classificazione-degli-stati-esteri/
ws = wb[wb.sheetnames[0]]

rows = ws.rows
next(rows)

for row in rows:
    if row[9].value in (None, 'n.d.'): continue
    territori[row[9].value] = (row[6].value.upper(), "EE")

print ('Carico gli Stati cessati...')
wb = load_workbook('Elenco-Paesi-esteri-cessati.xlsx') # https://www.istat.it/classificazione/classificazione-degli-stati-esteri/
ws = wb[wb.sheetnames[0]]

rows = ws.rows
next(rows)

for row in rows:
    if row[4].value in (None, 'n.d'): continue
    territori[row[4].value] = (row[7].value.upper(), "EE")

json.dump(territori, open('territori.json','w'))
print('Elenco dei codici catastali di Comuni e Stati esteri generato in territori.json')
