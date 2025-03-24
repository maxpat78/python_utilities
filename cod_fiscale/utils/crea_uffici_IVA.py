"Crea una tabella di codici ufficio IVA riportati da Wikipedia (alcuni mancano dall'Agenzia delle Entrate)"
from openpyxl import load_workbook
import json

print ('Carico i codici ufficio IVA...')
wb = load_workbook('uffici_IVA_wikipedia.xlsx')
ws = wb[wb.sheetnames[0]]

rows = ws.rows
next(rows)

uffici = {}
for row in rows:
    try:
        cod = '%03d' % int(row[0].value)
        uffici[cod] = row[1].value.upper()
    except:
        pass

json.dump(uffici, open('uffici.json','w'))
print('Elenco degli uffici IVA generato in uffici.json')
